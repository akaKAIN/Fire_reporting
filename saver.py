import openpyxl
import os
from messages.errors import ERRORS


class DocFile:
    def __init__(self):
        self.files_list = self.get_files_list()
        self.file_name = None
        self.active_sheet = None
        self.work_sheet = None
        self.workbook = None
        self.sheets_list = list()
        self.points_list = list()

    # Получение списка табличных файлов в текущей директории
    @staticmethod
    def get_files_list():
        path = os.getcwd()
        file_list = list()
        for elem in os.listdir(path):
            if os.path.isfile(elem) and elem.endswith((".xls", ".xlsx")):
                file_list.append(elem)
        file_list.sort()
        return file_list

    # Проверка существования файла. Чтение списка листов в файле
    def get_sheets_list(self, file_name: str) -> (list, str):
        if os.path.isfile(file_name):
            self.file_name = file_name
            self.workbook = openpyxl.load_workbook(file_name)
            self.sheets_list = self.workbook.sheetnames
            return self.sheets_list, ""
        return list(), ERRORS['no_file']

    def get_active_sheet(self, selected_sheet_name: str):
        if selected_sheet_name in self.sheets_list:
            self.active_sheet = selected_sheet_name
            self.work_sheet = self.workbook[self.active_sheet]

    def save_in_file(self, key, data):

        def addition_values(obj, value, row, column, aggregate=True):
            """Функция переноса данных в ячейки и агрегации переданных значений в Excel-файл"""

            # Сохраняем данные "за сутки"
            active_cell = obj.work_sheet.cell(row=row, column=column)
            active_cell.value = value
            if aggregate:
                # Сохраняем данные "за период", агрегируя "за сутки" с имеющимся значением "за период".
                active_cell = obj.work_sheet.cell(row=row, column=column + 1)

                # print(f'{next_cell.value=}\t{type(next_cell.value)=}')
                if active_cell.value is None:
                    active_cell.value = value
                else:
                    active_cell.value += value
            return

        cell = self.get_start_cell(key)

        # print(f'{key=}\n{cell=}\n{data=}')
        # Проверка существования выбраного пользователем имени подразделения в ячейке файла, куда сохраняем инфу.
        if getattr(cell, 'column', None) is None:
            print("не найдена ячейка с именем")
            return False, ERRORS["no_point"]

        for i in range(len(data) - 1):  # Последнее значение списка будет добавлено отдельно.
            # Сохраняем данные "за сутки"
            addition_values(self, value=data[i], row=cell.row, column=cell.column + 1 + (i * 2))

        # Добавление значения последней ячейки без последующей агрегации.
        addition_values(self, value=data[-1], row=cell.row, column=cell.column + len(data) * 2 - 1, aggregate=False)

        # TODO: удалять из заданной области название заполненной части (остаются только "незаполненные")
        self.workbook.save(self.file_name)
        return True, ""

    def get_start_cell(self, value_in_cell: str):
        if self.work_sheet:
            for cell_obj in self.work_sheet["B1:B300"]:
                for row in cell_obj:
                    if row.value == value_in_cell:
                        return row


def test(file_name):
    print(os.path.isfile(file_name))
    wb = openpyxl.load_workbook(file_name)
    sheet = wb['Лист1']


if __name__ == "__main__":
    print(os.getcwd())
    test('table.xlsx')
